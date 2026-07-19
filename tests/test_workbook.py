from pathlib import Path

from openpyxl import load_workbook

from prosway_acquisition.models import Company, SourceLog
from prosway_acquisition.workbook import ENTERPRISE_HEADERS, write_workbook, workbook_counts


def fake_company() -> Company:
    return Company(
        company_name="ACME",
        siren="123456789",
        segment_prosway="ESN/conseil IT",
        naf_code="62.01Z",
        naf_label="Programmation informatique",
        employee_range="100-199 salariés",
        city="Paris",
        region="Île-de-France",
        website="",
        source_company_url="https://annuaire-entreprises.data.gouv.fr/entreprise/123456789",
        fit_notes="notes",
        priority_score_0_100=95,
        priority_label="A",
        department="75",
    )


def test_company_as_enterprise_row_preserves_current_order():
    company = fake_company()
    assert company.as_enterprise_row() == [
        "ACME",
        "123456789",
        "ESN/conseil IT",
        "62.01Z",
        "Programmation informatique",
        "100-199 salariés",
        "Paris",
        "Île-de-France",
        "",
        "https://annuaire-entreprises.data.gouv.fr/entreprise/123456789",
        "notes",
        95,
        "A",
    ]
    assert ENTERPRISE_HEADERS == [
        "company_name",
        "siren",
        "segment_prosway",
        "naf_code",
        "naf_label",
        "employee_range",
        "city",
        "region",
        "website",
        "source_company_url",
        "fit_notes",
        "priority_score_0_100",
        "priority_label",
    ]


def test_workbook_generation_creates_expected_sheets_and_counts(tmp_path: Path):
    output = tmp_path / "prosway.xlsx"
    write_workbook(output, [fake_company()], [SourceLog("source", "query", "OK", "1 row")])

    wb = load_workbook(output, read_only=True, data_only=True)
    assert wb.sheetnames == ["Entreprises", "Signaux", "Contacts_cibles", "Scoring", "Sources_Logs"]
    assert workbook_counts(output) == {
        "Entreprises": 1,
        "Signaux": 1,
        "Contacts_cibles": 1,
        "Scoring": 6,
        "Sources_Logs": 4,
    }
