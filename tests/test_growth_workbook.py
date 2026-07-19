from pathlib import Path

from openpyxl import load_workbook

from prosway_acquisition.domain import CompanyProfile, Evidence, LeadResult
from prosway_acquisition.models import SourceLog
from prosway_acquisition.workbook import write_growth_workbook, workbook_counts


def test_growth_workbook_has_sales_action_sheets(tmp_path: Path):
    company = CompanyProfile("ACME", "123456789", employee_min=50, employee_max=99, employee_label="50-99 salariés", city="Paris", department="75", region="Île-de-France")
    lead = LeadResult(
        company=company,
        evidence=(Evidence(company.siren, "multi_hiring", "4 postes", "manual"),),
        priority_score_0_100=85,
        priority_label="A",
        recommended_angle="Structuration RH",
        enrichment_needed=False,
    )
    output = tmp_path / "growth.xlsx"
    write_growth_workbook(output, [lead], [SourceLog("source", "query", "OK")])
    wb = load_workbook(output, read_only=True, data_only=True)
    assert wb.sheetnames == ["Entreprises", "Signaux_croissance", "Angles_contact", "Scoring", "Sources_Logs"]
    assert workbook_counts(output)["Entreprises"] == 1
    assert workbook_counts(output)["Signaux_croissance"] == 1
