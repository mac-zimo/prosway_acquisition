from __future__ import annotations

import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from .config import TITLE

GOOGLE_WORKSPACE_SCRIPTS = Path.home() / ".hermes/skills/productivity/google-workspace/scripts"


def _load_google_api():
    if str(GOOGLE_WORKSPACE_SCRIPTS) not in sys.path:
        sys.path.insert(0, str(GOOGLE_WORKSPACE_SCRIPTS))
    import google_api  # type: ignore

    return google_api


def upload_as_google_sheet(xlsx_path: Path, title: str = TITLE) -> tuple[dict, object]:
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload

    google_api = _load_google_api()
    credentials = google_api.get_credentials()
    drive = build("drive", "v3", credentials=credentials)
    media = MediaFileUpload(
        str(xlsx_path),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=False,
    )
    metadata = {"name": title, "mimeType": "application/vnd.google-apps.spreadsheet"}
    file_info = drive.files().create(body=metadata, media_body=media, fields="id,name,mimeType,webViewLink").execute()
    return file_info, drive


def verify_google_sheet_export(drive: object, file_id: str) -> dict[str, object]:
    from googleapiclient.http import MediaIoBaseDownload

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    request = drive.files().export_media(
        fileId=file_id,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    downloader = MediaIoBaseDownload(tmp, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    tmp.close()

    try:
        wb = load_workbook(tmp.name, read_only=True, data_only=True)
        counts = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            counts[sheet_name] = max(0, sum(1 for _ in ws.iter_rows()) - 1)
        sample = None
        if "Entreprises" in wb.sheetnames:
            for row in wb["Entreprises"].iter_rows(min_row=2, max_row=2, values_only=True):
                sample = row[0]
                break
        return {"counts": counts, "sample_A2": sample, "sheetnames": wb.sheetnames}
    finally:
        Path(tmp.name).unlink(missing_ok=True)
