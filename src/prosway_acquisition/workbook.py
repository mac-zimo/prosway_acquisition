from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import ESN_NAF, INDUSTRY_NAF
from .models import Company, SourceLog
from .scoring import roles_for, signal_type_for

ENTERPRISE_HEADERS = [
    "company_name",
    "siren",
    "segment_prosway",
    "naf_code",
    "naf_label",
    "employee_range",
    "city",
    "region",
    "website",
    "source_company_url",
    "fit_notes",
    "priority_score_0_100",
    "priority_label",
]
SIGNAL_HEADERS = ["company_name", "siren", "signal_type", "signal_description", "source_url", "relevance_score_0_100"]
CONTACT_HEADERS = [
    "company_name",
    "siren",
    "target_roles_recommended",
    "public_contact_channel_found",
    "contact_source_url",
    "enrichment_needed",
    "notes",
]
LOG_HEADERS = ["Source/API/site", "Requête ou URL", "Statut", "Notes"]


def build_signal_rows(companies: list[Company]) -> list[list[object]]:
    rows = []
    for company in companies:
        description = (
            f"Entreprise active avec tranche d’effectif {company.employee_range} et code NAF "
            f"{company.naf_code} ({company.naf_label}) : base publique compatible avec prospection RH/QVCT/formation."
        )
        rows.append([
            company.company_name,
            company.siren,
            signal_type_for(company.segment_prosway),
            description,
            company.source_company_url,
            min(90, company.priority_score_0_100),
        ])
    return rows


def build_contact_rows(companies: list[Company]) -> list[list[object]]:
    return [
        [
            company.company_name,
            company.siren,
            roles_for(company.segment_prosway),
            "",
            "",
            "oui",
            "Pas de nom/email personnel collecté. Chercher uniquement canaux publics ou enrichissement conforme RGPD en phase 2.",
        ]
        for company in companies
    ]


def build_scoring_rows() -> list[list[str]]:
    return [
        ["Critère", "Poids / règle", "Commentaire"],
        ["Taille compatible 100-1000 salariés", "+15 à +20", "Tranches Sirene 22/31/32/41; cible Prosway prioritaire."],
        ["Segment prioritaire", "ESN/conseil IT +25; Industrie +22; Banque/assurance +12", "Segments MVP demandés en priorité; banque/assurance seulement opportuniste."],
        ["Source publique fiable", "+10", "SIREN/NAF/effectif issus de recherche-entreprises.api.gouv.fr / Annuaire des Entreprises."],
        ["Signal public RH/QVCT/HSE", "Score signal = min(90, priorité)", "Premier MVP: signal faible basé secteur+taille; signaux web fins à enrichir sans scraping agressif."],
        ["Labels", "A >=85; B 70-84; C 55-69; D <55", "Priorisation initiale, à recalibrer après validation commerciale."],
        ["Limite volontaire", "Sites web et contacts laissés vides si non trouvés via source publique simple", "Pas de LinkedIn scraping, pas d’Apollo/Kaspr/Dropcontact, pas d’invention."],
    ]


def build_log_rows(logs: list[SourceLog]) -> list[list[str]]:
    return [LOG_HEADERS, *[log.as_row() for log in logs],
        ["Hypothèse MVP", "Codes NAF ESN/conseil IT: " + ", ".join(ESN_NAF), "OK", "Cibles RH/coaching/transition/QVCT dans sociétés tech 100-1000."],
        ["Hypothèse MVP", "Codes NAF industrie échantillon: " + ", ".join(INDUSTRY_NAF), "OK", "Sélection non exhaustive de secteurs industriels avec potentiel RH/QVCT/HSE."],
        ["Contrainte RGPD", "Pas de collecte d’emails personnels; pas de scraping LinkedIn; pas d’outils payants", "OK", "Contacts_cibles contient des rôles recommandés et enrichment_needed=oui."],
    ]


def style_workbook(wb: Workbook) -> None:
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.alignment = Alignment(wrap_text=True)
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 22
        for wide_col in ("K", "L"):
            if ws.max_column >= ws[wide_col + "1"].column:
                ws.column_dimensions[wide_col].width = 60


def write_workbook(path: Path, companies: list[Company], logs: list[SourceLog]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    sheet_names = ["Entreprises", "Signaux", "Contacts_cibles", "Scoring", "Sources_Logs"]
    wb.active.title = sheet_names[0]
    for name in sheet_names[1:]:
        wb.create_sheet(name)

    wb["Entreprises"].append(ENTERPRISE_HEADERS)
    for company in companies:
        wb["Entreprises"].append(company.as_enterprise_row())

    wb["Signaux"].append(SIGNAL_HEADERS)
    for row in build_signal_rows(companies):
        wb["Signaux"].append(row)

    wb["Contacts_cibles"].append(CONTACT_HEADERS)
    for row in build_contact_rows(companies):
        wb["Contacts_cibles"].append(row)

    for row in build_scoring_rows():
        wb["Scoring"].append(row)

    for row in build_log_rows(logs):
        wb["Sources_Logs"].append(row)

    style_workbook(wb)
    wb.save(path)
    return path


def workbook_counts(path: Path) -> dict[str, int]:
    wb = load_workbook(path, read_only=True, data_only=True)
    counts = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        counts[sheet_name] = max(0, sum(1 for _ in ws.iter_rows()) - 1)
    return counts
