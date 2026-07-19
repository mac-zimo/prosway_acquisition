#!/usr/bin/env python3
import json, time, urllib.parse, urllib.request, sys, re, tempfile
from pathlib import Path
from datetime import datetime, timezone

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload

sys.path.insert(0, str(Path.home()/'.hermes/skills/productivity/google-workspace/scripts'))
import google_api

TITLE = 'Prosway - MVP Prospection RH'
OUT = Path.home()/'prosway_mvp_prospection_rh.xlsx'
LOG = []

NAF_LABELS = {
 '62.01Z':'Programmation informatique','62.02A':'Conseil en systèmes et logiciels informatiques','62.02B':'Tierce maintenance de systèmes et applications informatiques','62.03Z':'Gestion d’installations informatiques','62.09Z':'Autres activités informatiques','63.11Z':'Traitement de données, hébergement et activités connexes','70.22Z':'Conseil pour les affaires et autres conseils de gestion',
 '10.89Z':'Fabrication d’autres produits alimentaires n.c.a.','20.14Z':'Fabrication d’autres produits chimiques organiques de base','20.59Z':'Fabrication d’autres produits chimiques n.c.a.','22.29A':'Fabrication de pièces techniques à base de matières plastiques','25.62B':'Mécanique industrielle','26.51A':'Fabrication d’équipements d’aide à la navigation','27.12Z':'Fabrication de matériel de distribution et de commande électrique','28.29B':'Fabrication d’autres machines d’usage général','29.32Z':'Fabrication d’autres équipements automobiles','30.30Z':'Construction aéronautique et spatiale','33.12Z':'Réparation de machines et équipements mécaniques',
 '64.19Z':'Autres intermédiations monétaires','65.11Z':'Assurance vie','65.12Z':'Autres assurances','66.22Z':'Activités des agents et courtiers d’assurances'
}
ESN = ['62.01Z','62.02A','62.02B','62.03Z','62.09Z','63.11Z','70.22Z']
IND = ['10.89Z','20.14Z','20.59Z','22.29A','25.62B','26.51A','27.12Z','28.29B','29.32Z','30.30Z','33.12Z']
FIN = ['64.19Z','65.11Z','65.12Z','66.22Z']
TRANCHES = ['22','31','32','41']
TRANCHE_LABEL = {'22':'100-199 salariés','31':'200-249 salariés','32':'250-499 salariés','41':'500-999 salariés'}
DEPT_REGION = {
 **{str(i).zfill(2):'Île-de-France' for i in [75,77,78,91,92,93,94,95]},
 **{d:'Auvergne-Rhône-Alpes' for d in ['01','03','07','15','26','38','42','43','63','69','73','74']},
 **{d:'Bourgogne-Franche-Comté' for d in ['21','25','39','58','70','71','89','90']},
 **{d:'Bretagne' for d in ['22','29','35','56']},
 **{d:'Centre-Val de Loire' for d in ['18','28','36','37','41','45']},
 **{d:'Corse' for d in ['2A','2B']},
 **{d:'Grand Est' for d in ['08','10','51','52','54','55','57','67','68','88']},
 **{d:'Hauts-de-France' for d in ['02','59','60','62','80']},
 **{d:'Normandie' for d in ['14','27','50','61','76']},
 **{d:'Nouvelle-Aquitaine' for d in ['16','17','19','23','24','33','40','47','64','79','86','87']},
 **{d:'Occitanie' for d in ['09','11','12','30','31','32','34','46','48','65','66','81','82']},
 **{d:'Pays de la Loire' for d in ['44','49','53','72','85']},
 **{d:'Provence-Alpes-Côte d’Azur' for d in ['04','05','06','13','83','84']},
}

def api_get(params):
    url = 'https://recherche-entreprises.api.gouv.fr/search?' + urllib.parse.urlencode(params)
    req = urllib.request.Request(url, headers={'User-Agent':'Hermes Prosway MVP research'})
    try:
        with urllib.request.urlopen(req, timeout=20) as r:
            return json.load(r), url, None
    except Exception as e:
        return None, url, repr(e)

def city_from_siege(siege):
    return siege.get('libelle_commune') or siege.get('commune') or ''

def region_from_siege(siege):
    dep = siege.get('departement') or ''
    return DEPT_REGION.get(dep, '')

def segment_for(naf):
    if naf in ESN: return 'ESN/conseil IT'
    if naf in IND: return 'Industrie'
    if naf in FIN: return 'Banque/assurance'
    return 'Autre pertinent'

def annuaire_url(siren):
    return f'https://annuaire-entreprises.data.gouv.fr/entreprise/{siren}'

def score(segment, tranche, naf):
    s = 40
    s += {'ESN/conseil IT':25,'Industrie':22,'Banque/assurance':12}.get(segment,5)
    s += {'22':15,'31':18,'32':20,'41':18}.get(tranche,0)
    s += 10 # source API publique fiable + SIREN
    return min(100,s)

def label(s):
    return 'A' if s>=85 else 'B' if s>=70 else 'C' if s>=55 else 'D'

def collect():
    companies = {}
    target_per_naf = {naf: 10 for naf in ESN}
    target_per_naf.update({naf: 6 for naf in IND})
    target_per_naf.update({naf: 3 for naf in FIN})
    for naf in ESN + IND + FIN:
        got = 0
        for tr in TRANCHES:
            page = 1
            while got < target_per_naf[naf] and page <= 3:
                params = {'activite_principale': naf, 'tranche_effectif_salarie': tr, 'etat_administratif':'A', 'per_page':25, 'page':page}
                data, url, err = api_get(params)
                LOG.append(['API Recherche Entreprises', url, 'OK' if not err else 'ERREUR', err or f"{len(data.get('results', []))} résultats page"])
                if err or not data: break
                rows = data.get('results', [])
                if not rows: break
                for r in rows:
                    siren = r.get('siren')
                    if not siren or siren in companies: continue
                    siege = r.get('siege') or {}
                    if siege.get('etat_administratif') and siege.get('etat_administratif') != 'A': continue
                    eff = r.get('tranche_effectif_salarie') or siege.get('tranche_effectif_salarie') or tr
                    if eff not in TRANCHE_LABEL: continue
                    name = r.get('nom_complet') or r.get('nom_raison_sociale') or ''
                    seg = segment_for(naf)
                    sc = score(seg, eff, naf)
                    fit = f"Taille cible Prosway ({TRANCHE_LABEL.get(eff)}), segment {seg}, établissement siège actif d’après API publique. Site/contact à enrichir sans payant."
                    companies[siren] = {
                        'company_name': name[:120], 'siren': siren, 'segment_prosway': seg, 'naf_code': naf, 'naf_label': NAF_LABELS.get(naf,''),
                        'employee_range': TRANCHE_LABEL.get(eff,''), 'city': city_from_siege(siege), 'region': region_from_siege(siege),
                        'website':'', 'source_company_url': annuaire_url(siren), 'fit_notes': fit,
                        'priority_score_0_100': sc, 'priority_label': label(sc), '_tranche': eff
                    }
                    got += 1
                    if got >= target_per_naf[naf]: break
                page += 1
                time.sleep(0.15)
        if len(companies) >= 130: break
    return list(companies.values())

def roles_for(seg):
    base = 'DRH; RRH; Responsable Développement RH; Responsable Formation; HRBP; Talent/People Partner'
    if seg == 'Industrie':
        return base + '; Responsable QVCT/QVT; Responsable Prévention/HSE; Santé au travail'
    if seg == 'ESN/conseil IT':
        return base + '; Responsable Talent Acquisition; Learning & Development; Employee Experience'
    if seg == 'Banque/assurance':
        return base + '; Responsable QVCT; Responsable mobilité/carrières'
    return base

def build_rows(companies):
    entreprises = [[c[k] for k in ['company_name','siren','segment_prosway','naf_code','naf_label','employee_range','city','region','website','source_company_url','fit_notes','priority_score_0_100','priority_label']] for c in companies]
    signaux=[]; contacts=[]
    for c in companies:
        desc = f"Entreprise active avec tranche d’effectif {c['employee_range']} et code NAF {c['naf_code']} ({c['naf_label']}) : base publique compatible avec prospection RH/QVCT/formation."
        stype = 'HSE/QVCT potentiel' if c['segment_prosway']=='Industrie' else 'formation/recrutement potentiel' if c['segment_prosway']=='ESN/conseil IT' else 'développement RH potentiel'
        signaux.append([c['company_name'], c['siren'], stype, desc, c['source_company_url'], min(90, c['priority_score_0_100'])])
        contacts.append([c['company_name'], c['siren'], roles_for(c['segment_prosway']), '', '', 'oui', 'Pas de nom/email personnel collecté. Chercher uniquement canaux publics ou enrichissement conforme RGPD en phase 2.'])
    scoring = [
        ['Critère','Poids / règle','Commentaire'],
        ['Taille compatible 100-1000 salariés','+15 à +20','Tranches Sirene 22/31/32/41; cible Prosway prioritaire.'],
        ['Segment prioritaire','ESN/conseil IT +25; Industrie +22; Banque/assurance +12','Segments MVP demandés en priorité; banque/assurance seulement opportuniste.'],
        ['Source publique fiable','+10','SIREN/NAF/effectif issus de recherche-entreprises.api.gouv.fr / Annuaire des Entreprises.'],
        ['Signal public RH/QVCT/HSE','Score signal = min(90, priorité)','Premier MVP: signal faible basé secteur+taille; signaux web fins à enrichir sans scraping agressif.'],
        ['Labels','A >=85; B 70-84; C 55-69; D <55','Priorisation initiale, à recalibrer après validation commerciale.'],
        ['Limite volontaire','Sites web et contacts laissés vides si non trouvés via source publique simple','Pas de LinkedIn scraping, pas d’Apollo/Kaspr/Dropcontact, pas d’invention.'],
    ]
    logs = [['Source/API/site','Requête ou URL','Statut','Notes']] + LOG + [
        ['Hypothèse MVP','Codes NAF ESN/conseil IT: '+', '.join(ESN),'OK','Cibles RH/coaching/transition/QVCT dans sociétés tech 100-1000.'],
        ['Hypothèse MVP','Codes NAF industrie échantillon: '+', '.join(IND),'OK','Sélection non exhaustive de secteurs industriels avec potentiel RH/QVCT/HSE.'],
        ['Contrainte RGPD','Pas de collecte d’emails personnels; pas de scraping LinkedIn; pas d’outils payants','OK','Contacts_cibles contient des rôles recommandés et enrichment_needed=oui.'],
    ]
    return entreprises, signaux, contacts, scoring, logs

def make_xlsx(companies, rows):
    wb = Workbook()
    sheets = ['Entreprises','Signaux','Contacts_cibles','Scoring','Sources_Logs']
    wb.active.title = sheets[0]
    for s in sheets[1:]: wb.create_sheet(s)
    headers = {
        'Entreprises':['company_name','siren','segment_prosway','naf_code','naf_label','employee_range','city','region','website','source_company_url','fit_notes','priority_score_0_100','priority_label'],
        'Signaux':['company_name','siren','signal_type','signal_description','source_url','relevance_score_0_100'],
        'Contacts_cibles':['company_name','siren','target_roles_recommended','public_contact_channel_found','contact_source_url','enrichment_needed','notes'],
    }
    data_map = {'Entreprises': rows[0], 'Signaux': rows[1], 'Contacts_cibles': rows[2]}
    for name in ['Entreprises','Signaux','Contacts_cibles']:
        ws = wb[name]; ws.append(headers[name])
        for row in data_map[name]: ws.append(row)
    for row in rows[3]: wb['Scoring'].append(row)
    for row in rows[4]: wb['Sources_Logs'].append(row)
    for ws in wb.worksheets:
        ws.freeze_panes = 'A2'
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='1F4E79')
            cell.alignment = Alignment(wrap_text=True)
        for col in range(1, min(ws.max_column, 10)+1):
            ws.column_dimensions[get_column_letter(col)].width = 22
        if ws.max_column >= 10: ws.column_dimensions['K'].width = 60
    wb.save(OUT)

def upload_convert():
    creds = google_api.get_credentials()
    drive = build('drive','v3',credentials=creds)
    media = MediaFileUpload(str(OUT), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=False)
    metadata = {'name': TITLE, 'mimeType': 'application/vnd.google-apps.spreadsheet'}
    f = drive.files().create(body=metadata, media_body=media, fields='id,name,mimeType,webViewLink').execute()
    return f, drive

def verify(drive, file_id):
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    request = drive.files().export_media(fileId=file_id, mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    downloader = MediaIoBaseDownload(tmp, request)
    done = False
    while not done:
        status, done = downloader.next_chunk()
    tmp.close()
    wb = load_workbook(tmp.name, read_only=True, data_only=True)
    counts = {}
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        n = sum(1 for _ in ws.iter_rows())
        counts[ws_name] = max(0, n - 1)
    sample = None
    if 'Entreprises' in wb.sheetnames:
        for row in wb['Entreprises'].iter_rows(min_row=2, max_row=2, values_only=True):
            sample = row[0]
            break
    Path(tmp.name).unlink(missing_ok=True)
    return counts, sample, wb.sheetnames

if __name__ == '__main__':
    companies = collect()
    companies.sort(key=lambda c: (-c['priority_score_0_100'], c['segment_prosway'], c['company_name']))
    rows = build_rows(companies)
    make_xlsx(companies, rows)
    info, drive = upload_convert()
    counts, sample, sheetnames = verify(drive, info['id'])
    result = {'spreadsheetUrl': info.get('webViewLink'), 'id': info.get('id'), 'name': info.get('name'), 'counts': counts, 'sample_A2': sample, 'sheetnames': sheetnames, 'local_xlsx': str(OUT)}
    print(json.dumps(result, ensure_ascii=False, indent=2))
